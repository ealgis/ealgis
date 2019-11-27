#!/usr/bin/env python

#
# EAlGIS loader: Australian Census 2011; Data Pack 1
#

import re
import os
import glob
import os.path
import openpyxl
import sqlalchemy
import csv
import json
from datetime import datetime
from collections import OrderedDict

from ealgis.common.loaders import RewrittenCSV, CSVLoader
from ealgis.common.util import alistdir, make_logger
from .shapes import SHAPE_LINKAGE, SHAPE_SCHEMA
from .attrs_repair import repair_census_metadata, repair_column_series_census_metadata

logger = make_logger(__name__)


def parseColumnMetadata(table_number, column_name, metadata):
    """
    Parse Census DataPack column metadata into its component values:
        - Type (Row Name)
        - Kind (Column Name)
    """

    def getColumnLabel(kind):
        # Everything before the pipe is the column label, everything after is the series name
        # e.g. "Owner managers of unincorporated enterprises|MALES"
        # Column = Owner managers of unincorporated enterprises | Series = MALES
        return kind if "|" not in kind else kind.split("|")[0]

    def getSeriesName(kind):
        return None if "|" not in kind else kind.split("|")[1]

    def formatSeriesName(seriesName):
        return seriesName.replace(" - ", " ").replace("-", " ").replace("/ ", " ").replace("/", " ").replace("&", "and")

    def formatColumnLabel(columnLabel):
        """ Format the column label (kind) ready for parsing. """
        return columnLabel.strip().replace(": ", " ").replace(":", "").replace("-", " ").replace("$", "").replace(":", "").replace("\\", "").replace("&", "and").replace("/ ", " ").replace("/", " ").replace("etc.", "etc").replace(", ", " ")

    def formatHumanReadableRowLabel(rowLabel, rowType):
        rowLabel = rowLabel.replace("–", "-").strip()

        # Make currency ranges look nicer
        if "$" in rowType:
            match = re.search(
                r"(?P<rangeStart>[0-9]+)\s(?P<rangeEnd>[0-9]+)", rowLabel)
            if match is not None:
                rangeStart = "{:,}".format(int(match.group("rangeStart")))
                rangeEnd = "{:,}".format(int(match.group("rangeEnd")))
                rowLabel = "${}-${}".format(rangeStart, rangeEnd)

        return rowLabel

    def formatHumanReadableColumnLabel(columnLabel):
        return columnLabel.strip().replace("\\", "")

    metadata_original = metadata.copy()
    metadata = repair_census_metadata(table_number, column_name.lower(), metadata)
    metadata["type"] = metadata["type"].strip().replace("_", " ")

    columnType = metadata["type"].strip()
    columnLabel = formatColumnLabel(getColumnLabel(metadata["kind"]))
    seriesName = getSeriesName(metadata["kind"])

    # Special case for B02 and P02 - Selected Medians and Averages
    if table_number == "b02" or table_number == "p02":
        # No named columns here, just rows
        metadata["seriesName"] = None
        metadata["kind"] = ""

    elif seriesName is not None:
        # {SERIES NAME} {ROW LABEL} {COLUMN LABEL}
        # e.g. Persons Speaks other language and speaks English Total Year of arrival 2010
        # Series = Persons, Row = Speaks other language and speaks English Total, Column = Year of arrival 2010
        seriesName = formatSeriesName(seriesName)
        match = re.search(r"(?P<seriesName>{seriesName}) (?P<rowLabel>[A-z0-9\s]+) (?P<columnLabel>{columnLabel})".format(seriesName=seriesName, columnLabel=columnLabel), columnType, re.IGNORECASE)

        if match is not None:
            metadata["seriesName"] = seriesName
            metadata["type"] = formatHumanReadableRowLabel(match.group("rowLabel"), metadata["type"])
            metadata["kind"] = formatHumanReadableColumnLabel(getColumnLabel(metadata["kind"]))

    else:
        # {ROW LABEL} {COLUMN LABEL}
        # e.g. 150 299 Dwelling structure Flat unit or apartment In a 1 or 2 storey block
        # Row = 150 299, Column = Dwelling structure Flat unit or apartment In a 1 or 2 storey block
        match = re.search(r"(?P<rowLabel>[A-z0-9\s]+) (?P<columnLabel>{columnLabel})".format(columnLabel=columnLabel), columnType, re.IGNORECASE)
        if match is not None:
            metadata["seriesName"] = None
            metadata["type"] = formatHumanReadableRowLabel(match.group("rowLabel"), metadata["type"])
            metadata["kind"] = formatHumanReadableColumnLabel(metadata["kind"])

    if "seriesName" not in metadata:
        raise Exception("[{table_number}] Failed to parse column {column_name}: Fix '{kind}' // With '{type}'. Col label: '{columnLabel}'.".format(table_number=table_number, column_name=column_name, kind=metadata_original["kind"], type=metadata_original["type"].strip().replace("_", " "), columnLabel=columnLabel))

    # Discard - seriesName is now only used to validate parsing
    del metadata["seriesName"]

    return metadata


def load_metadata_table_serises(loader, census_dir, xlsx_name):
    """
    Parse Census metadata to extract the serises in each table.

    A series represents each set of data within a datapack, e.g.
    Males, Females, Persons

    Returns -
    col_meta[table_number][seriseName] = {
        "columns": [], # The Ids of the columns in a series.
        "datapackNames": [], # The names of the DataPack files (e.g. B12B, B12C) containing the columns for a series.
    }
    """

    def getSeriesName(kind):
        return None if "|" not in kind else kind.split("|")[1]

    col_meta = {}
    col_mapping = {}

    fname = os.path.join(census_dir + '/Metadata/', xlsx_name)
    logger.info("parsing metadata: %s" % (fname))
    wb = openpyxl.load_workbook(fname, read_only=True)

    def sheet_data(sheet):
        return (
            [t.value for t in r]
            for r in sheet.iter_rows()
            if len(r) > 0 and r[0].value is not None)

    def skip_to_descriptors(it):
        for row in sheet_iter:
            if row[0] != "Sequential":
                next(it)
            else:
                break

    sheet_iter = sheet_data(wb.worksheets[1])
    skip_to_descriptors(sheet_iter)
    for row in sheet_iter:
        name = row[0]
        if not name:
            continue
        name = name.lower()
        column_name, short_name, long_name, datapack_file, profile_table, column_heading = row[0:6]

        m = re.match('^([A-Za-z]+[0-9]+)([a-z]+)?$', datapack_file.lower())
        table_number = m.groups()[0]  # b46a -> b46

        column_heading = repair_column_series_census_metadata(table_number, column_name.lower(), str(column_heading).strip())
        seriseName = getSeriesName(column_heading)

        col_mapping[(table_number.lower(), column_name.lower())] = column_name

        if seriseName is not None:
            if table_number not in col_meta:
                col_meta[table_number] = {}

            if seriseName not in col_meta[table_number]:
                col_meta[table_number][seriseName] = {
                    "columns": [],
                    "datapackNames": [],
                }

            col_meta[table_number][seriseName]["columns"].append(column_name)

            if datapack_file.lower() not in col_meta[table_number][seriseName]["datapackNames"]:
                col_meta[table_number][seriseName]["datapackNames"].append(datapack_file.lower())
    del wb

    return col_meta, col_mapping


def load_metadata(loader, census_dir, xlsx_name, data_tables, columns_by_series, not_applicable_columns):
    table_meta = {}
    col_meta = {}

    fname = os.path.join(census_dir + '/Metadata/', xlsx_name)
    logger.info("parsing metadata: %s" % (fname))
    wb = openpyxl.load_workbook(fname, read_only=True)

    def sheet_data(sheet):
        return (
            [t.value for t in r]
            for r in sheet.iter_rows()
            if len(r) > 0 and r[0].value is not None)

    def skip(it, n):
        for i in range(n):
            next(it)

    def skip_to_descriptors(it):
        for row in sheet_iter:
            if row[0] != "Sequential":
                next(it)
            else:
                break

    def get_metadata_mapping():
        files = {}
        for json_file in glob.glob(os.path.join("./", "census2011", "*_metadata_mapping.json")):
            with open(json_file, "r") as f:
                files = {**files, **json.load(f)["tables"]}
        return files

    def get_topic_to_table_mapping():
        mapping = {}
        for json_file in glob.glob(os.path.join("./", "census2011", "*_topic_mapping.json")):
            with open(json_file, "r") as f:
                for topic_name, tables in json.load(f).items():
                    for table_number in tables:
                        table_number = table_number.upper()
                        if table_number not in mapping:
                            mapping[table_number] = []
                        mapping[table_number].append(topic_name)
        return mapping

    sheet_iter = sheet_data(wb.worksheets[0])
    skip(sheet_iter, 2)
    for row in sheet_iter:
        name = row[0]
        if not name:
            continue
        name = name.lower()
        table_meta[name] = {'type': row[1].strip(), 'kind': row[2].strip() if row[2] is not None else ""}

    sheet_iter = sheet_data(wb.worksheets[1])
    skip_to_descriptors(sheet_iter)
    for row in sheet_iter:
        name = row[0]
        if not name:
            continue
        name = name.lower()
        short_name, long_name, datapack_file, profile_table, column_heading = row[1:6]
        datapack_file = datapack_file.lower()
        m = re.match('^([A-Za-z]+[0-9]+)([a-z]+)?$', datapack_file)
        table_number = m.groups()[0]  # b46a -> b46
        if table_number not in col_meta:
            col_meta[table_number] = []

        try:
            meta = parseColumnMetadata(
                table_number,
                name,
                {'type': str(row[2]).strip(), 'kind': str(row[5]).strip()}
            )

            if name.lower() in not_applicable_columns:
                meta["na"] = True

            col_meta[table_number].append((name, meta))
        except Exception as e:
            if "object has no attribute" in str(e):
                print(name)
                raise e
            logger.error(e)
    del wb

    metadata_mapping = get_metadata_mapping()
    topic_to_table_mapping = get_topic_to_table_mapping()

    for table_name in data_tables:
        datapack_file = table_name.split('_', 1)[0].lower()
        m = re.match('^([A-Za-z]+[0-9]+)(s[0-9]{1,2})?$', datapack_file)
        table_number = m.groups()[0]  # b46a -> b46
        series_id = int(m.groups()[1][1:]) if m.groups()[1] is not None else None  # Just a number that increments from 1
        meta = table_meta[table_number]
        meta["series"] = None
        meta["family"] = table_number

        # Merge JSON formatted metadata
        if table_number.upper() in metadata_mapping:
            meta = {**meta, **metadata_mapping[table_number.upper()]}

        # Merge JSON formatted topic mappings
        if table_number.upper() in topic_to_table_mapping:
            meta["topics"] = topic_to_table_mapping[table_number.upper()]
        else:
            raise Exception("Couldn't find a topic mapping for table '%s'" % (table_number))

        columns = col_meta[table_number]

        if series_id is not None:
            if table_number not in columns_by_series:
                raise Exception("Expected to find serises for {}".format(table_number))

            series_name = list(columns_by_series[table_number].keys())[series_id - 1]
            meta["series"] = series_name

            # Filter all columns for the table down to just those columns in this series
            columns = [(col_name, col) for col_name, col in col_meta[table_number] if col_name.upper() in columns_by_series[table_number][series_name]["columns"]]

        # Validate metadata to ensure that we have the expected number
        # of rows and columns
        def isTableColumnMetadataValid(header, rows, column_uids):
            for h in header:
                for r in rows:
                    column_uid = "{}.{}".format(r, h)
                    if column_uid not in column_uids:
                        return False
            return True

        def getNumberOfMatchingHeaders(row, header, column_uids):
            count = 0
            headers = []
            for h in header:
                column_uid = "{}.{}".format(row, h)
                if column_uid in column_uids:
                    count += 1
                    headers.append(h)
            return count, headers

        header = list(set([c[1]["kind"] for c in columns]))
        rows = list(set([c[1]["type"] for c in columns]))
        column_uids = list(set(["{}.{}".format(c[1]["type"], c[1]["kind"]) for c in columns]))

        if not isTableColumnMetadataValid(header, rows, column_uids):
            logger.error("Table Header/Row mismatch found on table '{}' series '{}'".format(table_number, meta["series"]))

            # for r in rows:
            #     matches, headers = getNumberOfMatchingHeaders(r, header, column_uids)
            #     if matches != len(header):
            #         missing = [i for i in header if i not in headers]
            #         logger.error("Row '{}' has {} of {} matching headers. Missing: {}".format(r, matches, len(header), ", ".join(missing)))

        # print("#### header for series '{}'".format(meta["series"]), len(header))
        # for h in header:
        #     print(h)

        # print("#### rows for series '{}'".format(meta["series"]), len(rows))
        # for r in rows:
        #     print(r)

        # print("#### column_uids for series '{}'".format(meta["series"]), len(column_uids))
        # for uid in column_uids:
        #     print(uid)

        # print("#### columns for series '{}'".format(meta["series"]), len(columns))

        loader.set_table_metadata(table_name, meta)
        loader.register_columns(table_name, columns)


def load_datapacks(loader, census_dir, tmpdir, packname, abbrev, geo_gid_mapping, columns_by_series, col_mapping):
    def get_csv_files():
        files = []
        for geography in alistdir(d):
            # if "/SA3" in geography or "/SA1" in geography or "/LGA" in geography:
            logger.info("%s: Geograpy - %s" % (abbrev, geography))

            g = os.path.join(geography, "*.csv")
            csv_files = glob.glob(g)
            if len(csv_files) == 0:
                g = os.path.join(geography, "AUST", "*.csv")
                csv_files = glob.glob(g)
            if len(csv_files) == 0:
                raise Exception("can't find CSV files for `%s'" % geography)
            files += csv_files
        return files

    def get_csv_files_by_geography_and_table():
        csv_files = get_csv_files()
        by_table = {}

        for i, csv_path in enumerate(csv_files):
            if csv_path.endswith(".tmp.csv"):
                continue

            filename = os.path.basename(csv_path)
            datapack_file = filename.split('_', 1)[1].lower()
            m = re.match('^([A-Za-z]+[0-9]+)([a-z]+)?_.+$', datapack_file)
            table_number = m.groups()[0]
            geography_name = filename.split('_')[3].lower()

            if geography_name not in by_table:
                by_table[geography_name] = {}
            if table_number not in by_table[geography_name]:
                by_table[geography_name][table_number] = []
            by_table[geography_name][table_number].append(csv_path)
        return by_table

    def split_datapack_csv_by_series(columns_by_series, table_name, csv_path):
        csv_files = []

        for key, series_name in enumerate(columns_by_series[table_name]):
            with open(csv_path, "r") as merged_csv_file:
                # Open a new reader for each series as a means of resetting the pointer to the start of the file
                reader = csv.DictReader(merged_csv_file)

                series_csv_path = csv_path.replace(table_name.upper(), "{}S{}".format(table_name.upper(), key + 1))
                if not series_csv_path.endswith(".tmp.csv"):
                    series_csv_path = series_csv_path.replace(".csv", ".tmp.csv")

                # https://stackoverflow.com/a/39923823/7368493
                fieldnames = ["region_id"] + columns_by_series[table_name][series_name]["columns"]
                fieldnames_set = set(fieldnames)
                # logger.info("Fieldnames ({}): {}".format(len(fieldnames), fieldnames))

                with open(series_csv_path, "w") as f:
                    writer = csv.DictWriter(f, fieldnames)
                    writer.writeheader()

                    for row in reader:
                        # Use a dictionary comprehension to iterate over the key, value pairs
                        # discarding those pairs whose key is not in the set
                        filtered_row = dict(
                            (k, v) for k, v in row.items() if k in fieldnames_set
                        )
                        writer.writerow(filtered_row)

                logger.info("%s-%s: Created CSV file for series '%s' - %s" % (abbrev, table_name.upper(), series_name, os.path.basename(series_csv_path)))
                csv_files.append(series_csv_path)
        del reader

        return csv_files

    def split_datapack_csvs_by_series(columns_by_series, table_name, csv_paths):
        csv_files = []

        # logger.info("Start opening: {}".format(csv_path))
        # with open(csv_path, "r") as f:
        #     reader = list(csv.DictReader(f))
        # logger.info("Opened!")

        logger.info("Table Name: {}".format(table_name))
        logger.info("CSV Paths: {}".format([os.path.basename(path) for path in csv_paths]))
        # logger.info("CSV Paths 2: {}".format([os.path.basename(path).split("_")[1].lower() for path in csv_paths]))

        for key, series_name in enumerate(columns_by_series[table_name]):
            logger.info("")
            logger.info("Series Name: {}".format(series_name))

            # logger.info("0: {}".format(os.path.basename(csv_paths[0])))
            profiletable_name = os.path.basename(csv_paths[0]).split('_')[1]
            # logger.info("Profile Table Path Name: {}".format(profiletable_name))
            base_csv_path = csv_paths[0].replace("_{}_".format(profiletable_name), "_{}_".format(table_name.upper())).replace(".csv", ".tmp.csv")
            # logger.info("Merged Path: {}".format(os.path.basename(merged_csv_path)))
            series_csv_path = base_csv_path.replace(table_name.upper(), "{}S{}".format(table_name.upper(), key + 1))
            if not series_csv_path.endswith(".tmp.csv"):
                series_csv_path = series_csv_path.replace(".csv", ".tmp.csv")
            logger.info("Series CSV Path: {}".format(os.path.basename(series_csv_path)))

            # https://stackoverflow.com/a/39923823/7368493
            fieldnames = ["region_id"] + columns_by_series[table_name][series_name]["columns"]
            fieldnames_set = set(fieldnames)
            logger.info("Fieldnames ({}): {}".format(len(fieldnames), fieldnames))
            logger.info("Profile Tables: {}".format(columns_by_series[table_name][series_name]["datapackNames"]))

            with open(series_csv_path, "w") as f:
                writer = csv.DictWriter(f, fieldnames)
                writer.writeheader()

                # logger.info("Profile Table CSVs Test: {}".format([os.path.basename(path) for path in csv_paths if os.path.basename(path).split("_")[1].lower() in columns_by_series[table_name][series_name]["datapackNames"]]))
                profileTablesCSVPaths = [path for path in csv_paths if os.path.basename(path).split("_")[1].lower() in columns_by_series[table_name][series_name]["datapackNames"]]
                logger.info("Profile Table CSVs: {}".format([os.path.basename(path) for path in profileTablesCSVPaths]))

                for i, csv_path in enumerate(profileTablesCSVPaths):
                    logger.info("CSV File: {}".format(os.path.basename(csv_path)))

                    with open(csv_path, "r") as f:
                        reader = list(csv.DictReader(f))

                    for row in reader:
                        # Use a dictionary comprehension to iterate over the key, value pairs
                        # discarding those pairs whose key is not in the set
                        filtered_row = dict(
                            (k, v) for k, v in row.items() if k in fieldnames_set
                        )
                        # logger.info("Found: {}".format(len(filtered_row)))
                        writer.writerow(filtered_row)

            logger.info("%s-%s: Created CSV file for series '%s' - %s" % (abbrev, table_name.upper(), series_name, os.path.basename(series_csv_path)))
            csv_files.append(series_csv_path)

        return csv_files

    def merge_and_get_csv_files_by_table_and_series():
        csv_files_by_geog_and_table = get_csv_files_by_geography_and_table()
        csv_files = []

        for geography_name, tables in csv_files_by_geog_and_table.items():
            # if geography_name != "lga":
            #     continue

            for table_name, csv_paths in csv_files_by_geog_and_table[geography_name].items():
                # Merge the separate profile table/datapack CSVs into a single new  CSV file based on region_id (first column)
                # if table_name != "i10":
                #     continue

                if len(csv_paths) > 1:
                    dicts = []

                    for i, csv_path in enumerate(csv_paths):
                        with open(csv_path, "r") as f:
                            r = csv.reader(f)
                            if i == 0:
                                dicts.append(OrderedDict((row[0], row[1:]) for row in r))
                            else:
                                dicts.append({row[0]: row[1:] for row in r})

                    result = OrderedDict()
                    for d in tuple(dicts):
                        for key, value in d.items():
                            result.setdefault(key, []).extend(value)

                    profiletable_name = os.path.basename(csv_paths[0]).split('_')[1]
                    merged_csv_path = csv_paths[0].replace("_{}_".format(profiletable_name), "_{}_".format(table_name.upper())).replace(".csv", ".tmp.csv")
                    with open(merged_csv_path, "w") as f:
                        w = csv.writer(f)
                        for key, value in result.items():
                            w.writerow([key] + value)

                    # Some tables are large (and have multiple datapacks), but no serises (e.g. X03)
                    # For these tables we just merge into one combined CSV file...
                    if table_name not in columns_by_series:
                        logger.info("%s: Merged datapack CSV files - %s" % (abbrev, ", ".join([os.path.basename(i) for i in csv_paths])))
                        csv_files.append(merged_csv_path)
                    else:
                        # ...but others are large and DO have serises (e.g. X01)
                        # These we will also merge into one combined CSV file,
                        # then split our merged file into separate CSVs for each
                        # series in the datapack
                        split_csv_files = split_datapack_csv_by_series(columns_by_series, table_name, merged_csv_path)
                        logger.info("%s: Split multiple datapack CSV files - %s" % (abbrev, ", ".join([os.path.basename(i) for i in split_csv_files])))
                        csv_files += split_csv_files

                        # Remove temporary merged CSV path - we have individual CSV files for each series now
                        os.remove(merged_csv_path)

                else:
                    # Some tables are small enough to fit multiple serises in a single datapack CSV file (e.g. P05)
                    # So we need to split these into separate CSVs for each series too
                    if table_name in columns_by_series:
                        split_csv_files = split_datapack_csv_by_series(columns_by_series, table_name, csv_paths[0])
                        logger.info("%s: Split single datapack CSV file - %s" % (abbrev, ", ".join([os.path.basename(i) for i in split_csv_files])))
                        csv_files += split_csv_files
                    else:
                        csv_files.append(csv_paths[0])
        return csv_files

    def handleNotApplicableCells(value, column_name):
        """
        Detect cells that are 'Not Applicable' in the source data and
        set them to None (NULL in PostgreSQL).

        e.g. G23 has a row that refers to people who migrated to
        Australia before 2000, and a column that describes people who 
        are 14 years or younger. i.e. An impossibility.

        Cells that are not applicable have no value and need to be
        disabled in the Ealgis GUI so users can't select them.

        In the Census these are represented by the string ".."

        value (string): The value of a cell in a CSV file.
        column_name (string): The name of the column this cell is in e.g. g7068

        Returns:
            value (string or None)
        """
        # https://stackoverflow.com/a/23639915
        def is_number(s):
            """ Returns True is string is a number (int or float). """
            return s.replace('.', '', 1).isdigit()

        NotApplicableString = ".."

        if is_number(value) is False and value != NotApplicableString:
            logger.error("A cell contains an unknown value of \"{}\"".format(value))

        if column_name.lower() in not_applicable_columns and is_number(value) is True:
            # Remove column_name from stack if it looks like data
            not_applicable_columns.remove(column_name.lower())
        if column_name.lower() not in not_applicable_columns and value == NotApplicableString:
            # Add column_name to stack if it's Not Applicable
            not_applicable_columns.append(column_name.lower())

        return None if value == NotApplicableString else value

    d = os.path.join(census_dir, packname, "Sequential Number Descriptor")
    csv_files = merge_and_get_csv_files_by_table_and_series()

    table_re = re.compile(r'^2011Census_(.*)_sequential(.tmp)?.csv$')
    linkage_pending = []
    data_tables = []
    not_applicable_columns = []

    for i, csv_path in enumerate(csv_files):
        logger.info("%s: [%d/%d] %s" % (abbrev, i + 1, len(csv_files), os.path.basename(csv_path)))
        table_name = table_re.match(os.path.split(csv_path)[-1]).groups()[0].lower()

        m = re.match('^([a-z]+[0-9]+[a-z]{0,})(s[0-9]{1,2})?_.+', table_name.lower())
        table_number = m.groups()[0]  # g14_aus_ssc -> g14; g23s1_aus_ssc -> g23

        data_tables.append(table_name)
        decoded = table_name.split('_')

        if len(decoded) == 3:
            census_division = decoded[2]
        else:
            census_division = None

        gid_match = None

        if census_division is not None:
            def make_match_fn():
                lookup = geo_gid_mapping[census_division]
                header = None

                def _getColumnNameFromHeader(cell_index_in_row):
                    # We add 2 to the index because the first two columns are ["gid", "region_id"] and
                    # we're starting from position 1 in the row.
                    return header[cell_index_in_row + 2:cell_index_in_row + 3][0]

                def _matcher(line, row):
                    nonlocal header
                    if line == 0:
                        # Rewrite the header
                        # col_mapping: Map from ("G11", "Tot_P_M") (in the CSV header) to "G100" (in the database)
                        header = ["gid", "region_id"] + [col_mapping[(table_number, v.lower())] for v in row[1:]]
                        return header
                    else:
                        # Data rows
                        if row[0] in lookup:
                            return [str(lookup[row[0]])] + [row[0]] + [handleNotApplicableCells(v, _getColumnNameFromHeader(k)) for k, v in enumerate(row[1:])]
                        else:
                            # Fail dramatically if any missing gids have made it this far
                            raise Exception("failed gid lookup for '%s' for '%s'" % (row[0], census_division))
                return _matcher
            gid_match = make_match_fn()

        # normalise the CSV file by reading it in and writing it out again,
        # Postgres is quite pedantic. we also want to add an additional column to it
        with RewrittenCSV(tmpdir, csv_path, gid_match) as norm:
            instance = CSVLoader(loader.dbschema(), table_name, norm.get(), pkey_column=0)
            table_info = instance.load(loader)
            if table_info is not None and census_division is not None:
                linkage_pending.append((table_name, table_info, census_division))

        # Tidy up after ourselves
        if csv_path.endswith(".tmp.csv"):
            os.remove(csv_path)

    with loader.access_schema(SHAPE_SCHEMA) as geo_access:
        for attr_table, table_info, census_division in linkage_pending:
            geo_column, _, _ = SHAPE_LINKAGE[census_division]
            loader.add_geolinkage(
                geo_access,
                census_division, "gid",
                attr_table, "gid")

    return data_tables, not_applicable_columns


def build_geo_gid_mapping(factory):
    with factory.make_schema_access(SHAPE_SCHEMA) as shape_access:
        geo_gid_mapping = {}
        for census_division in SHAPE_LINKAGE:
            geo_column, geo_cast_required, _ = SHAPE_LINKAGE[census_division]
            geo_cls = shape_access.get_table_class(census_division, refresh=True)
            geo_attr = getattr(geo_cls, geo_column)
            if geo_cast_required is not None:
                inner_col = sqlalchemy.cast(geo_attr, geo_cast_required)
            else:
                inner_col = geo_attr
            lookup = {}
            for gid, match in shape_access.session.query(geo_cls.gid, inner_col).all():
                lookup[str(match)] = gid
            geo_gid_mapping[census_division] = lookup
        return geo_gid_mapping


def load_attrs(factory, census_dir, tmpdir):
    release = '3'

    packages = [
        ("Aboriginal and Torres Strait Islander Peoples Profile", "IP", "Metadata_2011_IP_DataPack.xlsx", "http://www.abs.gov.au/ausstats/abs@.nsf/papersbyReleaseDate/70B0E87BFC57CFE3CA257AA600136D3A?OpenDocument"),
        ("Basic Community Profile", "BCP", "Metadata_2011_BCP_DataPack.xlsx", "http://www.abs.gov.au/websitedbs/censushome.nsf/home/communityprofiles"),
        ("Place of Enumeration Profile", "PEP", "Metadata_2011_PEP_DataPack.xlsx", "http://www.abs.gov.au/ausstats/abs@.nsf/products/8862E7818AD89474CA2570D90018BFAF?OpenDocument"),
        ("Expanded Community Profile", "XCP", "Metadata_2011_XCP_DataPack.xlsx", "http://www.abs.gov.au/ausstats/abs@.nsf/mf/2069.0.30.005?OpenDocument"),
        ("Time Series Profile", "TSP", "Metadata_2011_TSP_DataPack.xlsx", "http://www.abs.gov.au/ausstats/abs@.nsf/ProductsbyReleaseDate/87541FA89DA17C6FCA257AA600136D72?OpenDocument"),
        ("Working Population Profile", "WPP", "Metadata_2011_WPP_DataPack.xlsx", "http://www.abs.gov.au/ausstats/abs@.nsf/productsbytitle/E6A94B5402FD62DCCA2570D90018BFAC?OpenDocument"),
    ]

    attr_results = []
    geo_gid_mapping = build_geo_gid_mapping(factory)
    for package_name, abbrev, metadata_filename, package_description in packages:
        dirname = '2011 ' + package_name + ' Release %s' % release
        schema_name = 'aus_census_2011_' + abbrev.lower()
        with factory.make_loader(schema_name) as loader:
            loader.add_dependency(SHAPE_SCHEMA)
            loader.set_metadata(
                name=package_name,
                family="ABS Census 2011",
                description=package_description,
                date_published=datetime(2012, 6, 21, 3, 0, 0)  # Set in UTC
            )
            columns_by_series, col_mapping = load_metadata_table_serises(loader, census_dir, metadata_filename)
            data_tables, not_applicable_columns = load_datapacks(loader, census_dir, tmpdir, dirname, abbrev, geo_gid_mapping, columns_by_series, col_mapping)
            load_metadata(loader, census_dir, metadata_filename, data_tables, columns_by_series, not_applicable_columns)
            attr_results.append(loader.result())
    return attr_results
