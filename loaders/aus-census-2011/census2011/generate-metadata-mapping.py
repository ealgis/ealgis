import os
import openpyxl
import re
import json


class ParserWarning(Exception):
    pass


def findFindOutMoreCell(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Find out more:":
                return cell
    # raise Exception("Failed to find the 'Find out more:' row.")
    return None


def findStartRow(sheet):
    findOutMoreCell = findFindOutMoreCell(sheet)
    if findOutMoreCell is None:
        return 2  # If there are metadata URLs (e.g. T01) then we start on row 2

    for row in sheet.iter_rows(row_offset=findOutMoreCell.row):
        cell = row[findOutMoreCell.col_idx - 1]
        if cell.value is None:
            return cell.row + 1
    raise Exception("Failed to find start row.")


def getMetadataURLs(sheet):
    findOutMoreCell = findFindOutMoreCell(sheet)
    if findOutMoreCell is None:
        return []  # No metadata URLs is OK (e.g. T01)

    metadataUrls = []
    for row in sheet.iter_rows(row_offset=findOutMoreCell.row):
        cell = row[findOutMoreCell.col_idx - 1]
        if cell.hyperlink is not None and cell.hyperlink.target is not None:
            metadataUrls.append({"name": cell.value, "url": cell.hyperlink.target})
    return metadataUrls


# def findSerises(sheet):
#     def findFirstColumnACellWithAValue(sheet):
#         for row in sheet.iter_rows(row_offset=findStartRow(sheet)):
#             if row[0].value is not None:
#                 return row[0].row
#         raise Exception("Failed to find the first data row.")

#     serises = []
#     for row in sheet.iter_rows(row_offset=findFirstColumnACellWithAValue(sheet) - 3):
#         hasValuesInColumnsOtherThanB = False
#         for cell in row:
#             # Check for empty strings to handle cases where columns contain empty strings (rather than None)
#             # or contain strings just containing whitespace.
#             if cell.column != "B" and cell.value is not None and str(cell.value).strip() != "":
#                 hasValuesInColumnsOtherThanB = True
#                 break
#         if hasValuesInColumnsOtherThanB is False and row[1].value is not None and row[1].value.isupper():
#             serises.append(row[1].value.strip())
#     return serises


def getNotes(sheet):
    def getFirstCellFromRow(row):
        if sheetName == "T 32d" or sheetName == "T 33" or\
                sheetName == "P 38b":
            return row[1]  # Column A in these tables is hidden...Because Reasons
        return row[0]

    def findNotesStartRow(sheet):
        firstRowWithAValue = None
        for row in reversed(list(sheet.iter_rows())):
            firstCell = getFirstCellFromRow(row)
            if firstRowWithAValue is None and firstCell.value is not None:
                # First row is also the only row with notes
                if str(firstCell.value).startswith("This table is based"):
                    return firstCell.row - 1
                firstRowWithAValue = firstCell.value
            elif firstRowWithAValue is not None:
                # The first row with no value signals the end of the notes section
                if firstCell.value is None or str(firstCell.value).strip() == "":
                    return firstCell.row
        raise Exception("Failed to find notes start row.")

    def findRowLabelsForNoteIdentifier(sheet, noteId):
        rowLabels = []
        for row in sheet.iter_rows(row_offset=findStartRow(sheet)):
            firstCell = getFirstCellFromRow(row)
            if firstCell.row == notesStartRow:
                break
            if firstCell.value is None:
                continue

            # print("Checking {noteId} in '{rowVal}'".format(noteId=noteId, rowVal=firstCell.value))
            match = re.search(r"^(?P<rowLabel>.+?)(?P<noteIdentifier>\({noteId}\))+[:]?".format(noteId=noteId), firstCell.value)
            if match is not None:
                # print("Found!")
                # print(match.groups(), "//", match.group("rowLabel").strip(), "//", match.group("noteIdentifier"))
                rowLabels.append(match.group("rowLabel").strip())
        # Where profile tables span multiple worksheets notes may refer to rows on another page (e.g. B01A)
        return rowLabels

    def linkNotesAndColumns(sheet, notes):
        formattedNotes = []
        for note in notes:
            match = re.search(r"^(?P<noteIdentifier>\([a-z]{1}\))\s", note)
            if match is not None:
                noteId = match.group("noteIdentifier")
                # print(noteId)

                rowLabels = findRowLabelsForNoteIdentifier(sheet, noteId)
                if len(rowLabels) > 0:
                    # print(rowLabels)
                    # Reformat the note to include the name of the row we're referring to e.g.
                    # "(a) Applicable to persons who are of both Aboriginal and Torres Strait Islander origin."
                    # becomes
                    # "<span class="rowLabel">Both Aboriginal and Torres Strait Islander</span> -  Applicable to persons who are of both Aboriginal and Torres Strait Islander origin."

                    formattedNotes.append("<span class=\"rowLabel\">{rowLabel}</span> - {note}".format(rowLabel="; ".join(rowLabels), note=note.replace(noteId, "")))
                else:
                    formattedNotes.append(note)
                    # raise ParserWarning("Warning: Could not find a matching row for note.")
            else:
                # For the start of the notes e.g. "This table is based on..."
                formattedNotes.append(note)
            # print
        return formattedNotes

    # The whole ERP series has two tables and no notes in either
    if sheetName == "E01" or sheetName == "E02":
        return []

    notesStartRow = findNotesStartRow(sheet)
    # print("notesStartRow", notesStartRow)

    notes = []
    for row in sheet.iter_rows(row_offset=notesStartRow):
        firstCell = getFirstCellFromRow(row)
        if firstCell.value is None:
            break
        notes.append(firstCell.value.strip())

    notes = linkNotesAndColumns(sheet, notes)
    return notes


def mergeNotesFromMultipleProfileTables(notes, newNotes):
    mergedNotes = []
    for key, note in enumerate(notes):
        if note.startswith("This table is based") or note.startswith("<span "):
            mergedNotes.append(note)
        else:
            if newNotes[key].startswith("<span "):
                mergedNotes.append(newNotes[key])
            else:
                mergedNotes.append(note)
    return mergedNotes


# def validateMetadataMapping(datapackMetadataFile, metadataMapping):
#     def formatRowLabel(rowLabel):
#         # Dependent_children_aged_21_24_years_female_parent_Total_male_parent_Language_and_proficiency_in_English_not_stated
#         return rowLabel.replace(" ", "_").replace("-", "_").replace(":", "")

#     def formatSeriesName(seriesName):
#         return seriesName.replace(" ", "_").replace("-", "_")

#     def formatCellHeading(cellHeading):
#         return cellHeading.replace(" ", "_").replace("-", "_").replace(":", "")

#     xl = openpyxl.load_workbook(os.path.join(base_dir, datapackMetadataFile), read_only=True)
#     cellDescriptorWorksheet = xl["Cell descriptors information"]

#     columnMapping = {}
#     for row in cellDescriptorWorksheet.iter_rows(row_offset=4):
#         longDescriptor = row[2].value
#         dataPack = row[3].value
#         profileTable = row[4].value
#         heading = row[5].value

#         if profileTable not in columnMapping:
#             columnMapping[profileTable] = []

#         columnMapping[profileTable].append({
#             "longDescriptor": longDescriptor,
#             "dataPack": dataPack,
#             "profileTable": profileTable,
#             "heading": heading,
#         })

#     for table in metadataMapping["tables"]:
#         # if "B13a" in table["profileTableNames"]:
#         #     continue
#         print(", ".join(table["profileTableNames"]))

#         for profileTableName in table["profileTableNames"]:
#             if profileTableName not in columnMapping or len(columnMapping[profileTableName]) == 0:
#                 raise ParserWarning("Unable to find any matching columns for profile table '{}'.".format(profileTableName))
#                 continue

#             if len(table["serises"]) == 0:
#                 numberOfRows = len([r for r in columnMapping[profileTableName]])
#                 if numberOfRows > 0:
#                     print("Rows: {}".format(numberOfRows))
#                 else:
#                     raise ParserWarning("Could not find any matching columns for tables '{}'.".format(", ".join(table["profileTableNames"])))

#         for seriesName in table["serises"]:
#             numberOfRows = 0
#             for profileTableName in table["profileTableNames"]:
#                 numberOfRows = len([r for r in columnMapping[profileTableName] if r["heading"].endswith("|{}".format(seriesName))])
#                 if numberOfRows > 0:
#                     break
#             if numberOfRows > 0:
#                 print("{}: {}".format(seriesName, numberOfRows))
#             else:
#                 raise ParserWarning("Could not find any matching columns for series '{}' in tables '{}'.".format(seriesName, ", ".join(table["profileTableNames"])))

#         print


base_dir = "/Users/keithmoss/Documents/Work/GitHub/ealgis/data-loader/data/ealgis-aus-census-2011-master/2011 Datapacks BCP_IP_TSP_PEP_ECP_WPP_ERP_Release 3_Partial/Metadata/"
# seriesName = "MALES"

# xl = openpyxl.load_workbook(os.path.join(base_dir, "Profile_template_with sequential_numbers_2011_BCP.xlsx"))
# xl = openpyxl.load_workbook(os.path.join(base_dir, "Profile_template_with sequential_numbers_2011_IP.xlsx"))
# xl = openpyxl.load_workbook(os.path.join(base_dir, "Profile_template_with sequential_numbers_2011_TSP.xlsx"))
# xl = openpyxl.load_workbook(os.path.join(base_dir, "Profile_template_with_sequential_numbers_2011_ERP.xlsx"))
# xl = openpyxl.load_workbook(os.path.join(base_dir, "Profile_template_with_sequential_numbers_2011_PEP.xlsx"))
# xl = openpyxl.load_workbook(os.path.join(base_dir, "Profile_template_with_sequential_numbers_2011_WPP.xlsx"))
xl = openpyxl.load_workbook(os.path.join(base_dir, "Profile_template_with_sequential_numbers_2011_XCP.xlsx"))
# print(xl.get_sheet_names())


metadataMapping = {"tables": {}}

for sheetName in xl.get_sheet_names()[1:]:
    # if sheetName == "B 04" or sheetName == "B 02":
    #     continue
    # if not (sheetName == "I 01a" or sheetName == "I 01b"):
    #     continue
    # if sheetName != "T 32d":
    #     continue
    m = re.match('^([A-Za-z]+[0-9]+)([a-z]+)?$', sheetName.replace(" ", ""))
    table_number = m.groups()[0]  # b46a -> b46
    print("{} ({})".format(sheetName, table_number))

    sheet = xl[sheetName]

    # serises = findSerises(sheet)
    metadataUrls = getMetadataURLs(sheet)
    notes = getNotes(sheet)

    if table_number not in metadataMapping["tables"]:
        metadataMapping["tables"][table_number] = {
            "metadataUrls": metadataUrls,
            "notes": notes,
        }
    else:
        metadataMapping["tables"][table_number]["notes"] = mergeNotesFromMultipleProfileTables(metadataMapping["tables"][table_number]["notes"], notes)

for table_name in metadataMapping["tables"]:
    metadataMapping["tables"][table_name]["notes"] = "<br />".join(metadataMapping["tables"][table_name]["notes"])

with open("xcp_metadata_mapping.json", "w") as f:
    f.write(json.dumps(metadataMapping, indent=2, sort_keys=True))


# Validate Metadata Mapping
# with open("ip_metadata_mapping.json", "r") as f:
#     metadataMapping = json.loads(f.read())

# validateMetadataMapping("Metadata_2011_BCP_DataPack.xlsx", metadataMapping)
# validateMetadataMapping("Metadata_2011_IP_DataPack.xlsx", metadataMapping)


# Basic Community Profile
# Merged: B12a-c, B17a-b, B23a-b, B40a-b, B41a-b, B42a-b, B43a-c
# Special Cases (Merged):
# B01a-b. Merged profileTableNames and notes by-hand
# B02. Manually created record and compiled notes.
# B04. Manually created record and compiled notes.
# B10a-b. Merged profileTableNames. Took notes from B10a.
# B13a-b. Merged profileTableNames and notes by-hand. Manually set profileTableName to "B13" to match what's in the metadata.

# Indigenous Profile
