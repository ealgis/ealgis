import os
import openpyxl
import re
import json


class ParserWarning(Exception):
    pass


def findFindOutMoreCell(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Find out more:":
                return cell
    # raise Exception("Failed to find the 'Find out more:' row.")
    return None


def findStartRow(sheet):
    findOutMoreCell = findFindOutMoreCell(sheet)
    if findOutMoreCell is None:
        return 2  # If there are metadata URLs (e.g. T01) then we start on row 2

    for row in sheet.iter_rows(row_offset=findOutMoreCell.row):
        cell = row[findOutMoreCell.col_idx - 1]
        if cell.value is None:
            return cell.row + 1
    raise Exception("Failed to find start row.")


def getMetadataURLs(sheet):
    findOutMoreCell = findFindOutMoreCell(sheet)
    if findOutMoreCell is None:
        return []  # No metadata URLs is OK (e.g. T01)

    metadataUrls = []
    for row in sheet.iter_rows(row_offset=findOutMoreCell.row):
        cell = row[findOutMoreCell.col_idx - 1]
        # print("cell.value", cell.value, cell)
        if cell.hyperlink is not None and cell.hyperlink.target is not None:
            # print("Add", cell.value)
            metadataUrls.append({"name": cell.value, "url": cell.hyperlink.target})
        else:
            # Links with long names sometimes span two cells -
            # so they're actually in the column to the right of this one.
            cell = row[findOutMoreCell.col_idx - 2]
            # print("cell.value", cell.value)
            if cell.hyperlink is not None and cell.hyperlink.target is not None:
                # print("Add", cell.value)
                metadataUrls.append({"name": cell.value, "url": cell.hyperlink.target})
    return metadataUrls


def getNotes(sheet):
    def getFirstCellFromRow(row):
        # if sheetName == "T 32d" or sheetName == "T 33" or\
        #         sheetName == "P 38b":
        #     return row[1]  # Column A in these tables is hidden...Because Reasons
        return row[0]

    def findNotesStartRow(sheet):
        firstRowWithAValue = None
        for row in reversed(list(sheet.iter_rows())):
            firstCell = getFirstCellFromRow(row)
            # print("firstCell", firstCell, firstCell.value)
            if firstRowWithAValue is None and firstCell.value is not None:
                # First row is also the only row with notes
                # print("firstRowWithAValue")
                if str(firstCell.value).startswith("This table is based"):
                    # print("foo 1")
                    return firstCell.row - 1
            #     firstRowWithAValue = firstCell.value
            # elif firstRowWithAValue is not None:
            #     # The first row with no value signals the end of the notes section
            #     if firstCell.value is None or str(firstCell.value).strip() == "":
            #         # print("foo 2")
            #         return firstCell.row
        raise Exception("Failed to find notes start row.")

    def findRowLabelsForNoteIdentifier(sheet, noteId):
        rowLabels = []
        for row in sheet.iter_rows(row_offset=findStartRow(sheet)):
            firstCell = getFirstCellFromRow(row)
            if firstCell.row == notesStartRow:
                break
            if firstCell.value is None:
                continue

            # print("Checking {noteId} in '{rowVal}' (cell {cell})".format(noteId=noteId, rowVal=firstCell.value, cell=firstCell))
            # Coerce firstCell's value to string to deal with row labels that
            # are purely numbers (e.g. ages)
            match = re.search(r"^(?P<rowLabel>.+?)(?P<noteIdentifier>\({noteId}\))+[:]?".format(noteId=noteId), str(firstCell.value))
            if match is not None:
                # print("Found!")
                # print(match.groups(), "//", match.group("rowLabel").strip(), "//", match.group("noteIdentifier"))
                rowLabels.append(match.group("rowLabel").strip())
        # Where profile tables span multiple worksheets notes may refer to rows on another page (e.g. B01A)
        return rowLabels

    def linkNotesAndColumns(sheet, notes):
        formattedNotes = []
        for note in notes:
            match = re.search(r"^(?P<noteIdentifier>\([a-z]{1}\))\s", note)
            if match is not None:
                noteId = match.group("noteIdentifier")
                # print(noteId)

                rowLabels = findRowLabelsForNoteIdentifier(sheet, noteId)
                if len(rowLabels) > 0:
                    # print(rowLabels)
                    # Reformat the note to include the name of the row we're referring to e.g.
                    # "(a) Applicable to persons who are of both Aboriginal and Torres Strait Islander origin."
                    # becomes
                    # "<span class="rowLabel">Both Aboriginal and Torres Strait Islander</span> -  Applicable to persons who are of both Aboriginal and Torres Strait Islander origin."

                    formattedNotes.append("<span class=\"rowLabel\">{rowLabel}</span> - {note}".format(rowLabel="; ".join(rowLabels), note=note.replace(noteId, "")))
                else:
                    formattedNotes.append(note)
                    # raise ParserWarning("Warning: Could not find a matching row for note.")
            else:
                # For the start of the notes e.g. "This table is based on..."
                formattedNotes.append(note)
            # print
        return formattedNotes

    # The Selected Medians and Averages tables have no "This table is based" line that
    # we can use to identify where the notes start - so we hard code it.
    if sheetName == "G 02":  # GCP
        notesStartRow = 23
    elif sheetName == "I 04":  # ATSIP
        notesStartRow = 28
    elif sheetName == "T 02":  # TSP
        notesStartRow = 21
    else:
        notesStartRow = findNotesStartRow(sheet)
    # print("notesStartRow", notesStartRow)

    notes = []
    for row in sheet.iter_rows(row_offset=notesStartRow):
        firstCell = getFirstCellFromRow(row)
        if firstCell.value is not None:
            notes.append(firstCell.value.strip())

            if firstCell.value.startswith("Please note that there are small random adjustments"):
                break

    notes = linkNotesAndColumns(sheet, notes)
    return notes


def mergeNotesFromMultipleProfileTables(notes, newNotes):
    mergedNotes = []
    for key, note in enumerate(notes):
        if note.startswith("This table is based") or note.startswith("<span "):
            mergedNotes.append(note)
        else:
            if newNotes[key].startswith("<span "):
                mergedNotes.append(newNotes[key])
            else:
                mergedNotes.append(note)
    return mergedNotes


base_dir = "/app/data/2016 Datapacks"

TABLES = [
    ["GCP", "2016 General Community Profile", "2016_GCP_Sequential_Template.xlsx"],
    ["PEP", "2016 Place of Enumeration Profile", "2016_PEP_Sequential_Template.xlsx"],
    ["ATSIP", "2016 Aboriginal and Torres Strait Islander Peoples Profile", "2016_ATSIP_Sequential_Template.xlsx"],
    ["TSP", "2016 Time Series Profile", "2016_TSP_Sequential_Template.xlsx"],
    ["WPP", "2016 Working Population Profile", "2016_WPP_Sequential_Template.xlsx"],
]

for table in TABLES:
    abbreviation = table[0]
    package_name = table[1]
    xls_name = table[2]

    xl = openpyxl.load_workbook(os.path.join(base_dir, package_name, "Metadata", xls_name))

    metadataMapping = {"tables": {}}

    for sheetName in xl.get_sheet_names():
        # if sheetName == "B 04" or sheetName == "B 02":
        #     continue
        # if not (sheetName == "I 01a" or sheetName == "I 01b"):
        #     continue
        # if sheetName == "G 02":
        #     continue
        # if sheetName != "G 03":
        #     continue
        # print("sheetName", sheetName)

        m = re.match('^([A-Za-z]+[0-9]+)([a-z]+)?$', sheetName.replace(" ", ""))
        if m is None:
            print("Skipping worksheet", sheetName)
            continue

        table_number = m.groups()[0]  # b46a -> b46
        print("{} ({})".format(sheetName, table_number))

        sheet = xl[sheetName]

        metadataUrls = getMetadataURLs(sheet)
        notes = getNotes(sheet)

        if table_number not in metadataMapping["tables"]:
            metadataMapping["tables"][table_number] = {
                "metadataUrls": metadataUrls,
                "notes": notes,
            }
        else:
            metadataMapping["tables"][table_number]["notes"] = mergeNotesFromMultipleProfileTables(metadataMapping["tables"][table_number]["notes"], notes)

    for table_name in metadataMapping["tables"]:
        metadataMapping["tables"][table_name]["notes"] = "<br />".join(metadataMapping["tables"][table_name]["notes"])

    json_path = "metadata_mappings/{}_metadata_mapping.json".format(abbreviation.lower())
    with open(json_path, "w") as f:
        f.write(json.dumps(metadataMapping, indent=2, sort_keys=True))
